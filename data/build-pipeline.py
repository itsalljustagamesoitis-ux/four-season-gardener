#!/usr/bin/env python3
"""
Build data/pipeline.json from the launch-200 Excel spreadsheet.
Resolves hub_slug, hub_label, category_slug, category_label from navigation.yaml.
Products list is empty — filled by the assignment pass.
"""

import json
import yaml
import openpyxl
from pathlib import Path

EXCEL_PATH = Path.home() / "Downloads/garden-launch-200-remapped.xlsx"
NAV_PATH = Path(__file__).parent.parent / "config/navigation.yaml"
OUT_PATH = Path(__file__).parent / "pipeline.json"

# cluster slug → (hub_slug, hub_label, category_slug, category_label)
def build_cluster_map(nav: dict) -> dict:
    mapping = {}
    for cat in nav["categories"]:
        for hub in cat.get("hubs", []):
            mapping[hub["slug"]] = {
                "hub_slug": hub["slug"],
                "hub_label": hub["label"],
                "category_slug": cat["slug"],
                "category_label": cat["label"],
            }
    return mapping

def main():
    with open(NAV_PATH) as f:
        nav = yaml.safe_load(f)

    cluster_map = build_cluster_map(nav)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Launch-200 Remapped"]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    pipeline = []
    for row in range(2, ws.max_row + 1):
        num = ws.cell(row, col["#"]).value
        if not num:
            continue

        cluster = ws.cell(row, col["Cluster"]).value or ""
        cluster = cluster.strip()

        nav_info = cluster_map.get(cluster, {
            "hub_slug": cluster,
            "hub_label": "",
            "category_slug": "",
            "category_label": "",
        })

        article = {
            "id": int(num),
            "keyword": ws.cell(row, col["Keyword"]).value or "",
            "slug": ws.cell(row, col["Suggested URL slug"]).value or "",
            "type": ws.cell(row, col["Article Type"]).value or "",
            "layout": ws.cell(row, col["Layout Type"]).value or "",
            "cluster": cluster,
            "hub_slug": nav_info["hub_slug"],
            "hub_label": nav_info["hub_label"],
            "hub_url": f"/{nav_info['hub_slug']}/",
            "category_slug": nav_info["category_slug"],
            "category_label": nav_info["category_label"],
            "category_url": f"/{nav_info['category_slug']}/",
            "angle": ws.cell(row, col["Angle / persona hook"]).value or "",
            "volume": ws.cell(row, col["Volume"]).value or 0,
            "kd": ws.cell(row, col["KD"]).value or 0,
            "h2_structure": ws.cell(row, col["H2 Structure"]).value or "",
            "products": [],
            "assignment_notes": "",
            "status": "pending",
        }
        pipeline.append(article)

    with open(OUT_PATH, "w") as f:
        json.dump(pipeline, f, indent=2)

    print(f"Written {len(pipeline)} articles to {OUT_PATH}")

    # sanity check clusters
    found = {a["cluster"] for a in pipeline}
    missing = found - set(cluster_map.keys())
    if missing:
        print(f"WARNING: clusters not in navigation.yaml: {missing}")
    else:
        print("All clusters resolved from navigation.yaml.")

if __name__ == "__main__":
    main()
